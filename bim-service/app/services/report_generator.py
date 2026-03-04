"""Excel report generation for BIM data."""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import Room, Area, Material, Project, QuantityItem


# DCAB brand colors
HEADER_FILL = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1B3A5C")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="1B3A5C")
SUM_FILL = PatternFill(start_color="E8EDF2", end_color="E8EDF2", fill_type="solid")
SUM_FONT = Font(name="Calibri", bold=True, size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def _auto_column_width(ws) -> None:
    """Auto-fit column widths based on content."""
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 3, 45)


def _write_header(ws, row: int, headers: list[str]) -> int:
    """Write styled header row."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    return row + 1


def _write_row(ws, row: int, values: list, number_cols: set[int] | None = None) -> int:
    """Write a data row with styling."""
    number_cols = number_cols or set()
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        if col in number_cols:
            cell.number_format = '#,##0.0'
            cell.alignment = Alignment(horizontal="right")
    return row + 1


def _write_sum_row(ws, row: int, label: str, col_count: int, sum_col: int, sum_value: float) -> int:
    """Write a summary/total row."""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = SUM_FILL
        cell.font = SUM_FONT
        cell.border = THIN_BORDER

    ws.cell(row=row, column=1, value=label)
    sum_cell = ws.cell(row=row, column=sum_col, value=sum_value)
    sum_cell.number_format = '#,##0.0'
    sum_cell.alignment = Alignment(horizontal="right")
    return row + 1


def generate_raumbuch(project: Project, rooms: list[Room]) -> io.BytesIO:
    """Generate Raumbuch (room book) Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Raumbuch"

    # Title
    ws.cell(row=1, column=1, value=f"Raumbuch - {project.name}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=project.address).font = SUBTITLE_FONT
    ws.cell(row=3, column=1, value=f"Stand: {project.status}").font = BODY_FONT

    headers = ["Nr.", "Raumname", "Geschoss", "Fläche (m²)", "Höhe (m)", "Volumen (m³)",
               "Nutzungsart", "Boden", "Wand", "Decke"]
    row = _write_header(ws, 5, headers)

    number_cols = {4, 5, 6}
    current_floor = None
    floor_area = 0.0

    for room in sorted(rooms, key=lambda r: (r.floor, r.number)):
        if current_floor is not None and room.floor != current_floor:
            row = _write_sum_row(ws, row, f"Summe {current_floor}", len(headers), 4, floor_area)
            floor_area = 0.0
            row += 1  # empty row

        current_floor = room.floor
        floor_area += room.area

        row = _write_row(ws, row, [
            room.number, room.name, room.floor, room.area, room.height, room.volume,
            room.usage_type, room.finish_floor, room.finish_wall, room.finish_ceiling,
        ], number_cols)

    # Final floor sum
    if current_floor is not None:
        row = _write_sum_row(ws, row, f"Summe {current_floor}", len(headers), 4, floor_area)

    # Grand total
    row += 1
    total_area = sum(r.area for r in rooms)
    row = _write_sum_row(ws, row, "GESAMT", len(headers), 4, total_area)

    _auto_column_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_flaechen(project: Project, areas: list[Area]) -> io.BytesIO:
    """Generate Flächenübersicht (area overview) Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Flächenübersicht"

    ws.cell(row=1, column=1, value=f"Flächenübersicht DIN 277 - {project.name}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=project.address).font = SUBTITLE_FONT

    headers = ["Bezeichnung", "Kategorie", "Geschoss", "Fläche (m²)", "Anzahl Räume"]
    row = _write_header(ws, 4, headers)

    number_cols = {4, 5}
    cat_totals: dict[str, float] = {}

    for area in sorted(areas, key=lambda a: (a.category, a.floor)):
        cat_totals[area.category] = cat_totals.get(area.category, 0) + area.area
        row = _write_row(ws, row, [
            area.name, area.category, area.floor, area.area, len(area.rooms),
        ], number_cols)

    row += 1
    for cat, total in sorted(cat_totals.items()):
        cat_labels = {"NUF": "Nutzungsfläche (NUF)", "TF": "Technische Funktionsfläche (TF)", "VF": "Verkehrsfläche (VF)"}
        row = _write_sum_row(ws, row, cat_labels.get(cat, cat), len(headers), 4, total)

    row += 1
    row = _write_sum_row(ws, row, "Netto-Raumfläche (NRF)", len(headers), 4, sum(cat_totals.values()))

    _auto_column_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def generate_materialien(project: Project, materials: list[Material]) -> io.BytesIO:
    """Generate Materialliste (material list) Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Materialliste"

    ws.cell(row=1, column=1, value=f"Materialliste - {project.name}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=project.address).font = SUBTITLE_FONT

    headers = ["Material", "Kategorie", "Menge", "Einheit", "Hersteller", "Produkt"]
    row = _write_header(ws, 4, headers)

    number_cols = {3}
    current_cat = None

    for mat in sorted(materials, key=lambda m: (m.category, m.name)):
        if current_cat is not None and mat.category != current_cat:
            row += 1  # visual separator
        current_cat = mat.category

        row = _write_row(ws, row, [
            mat.name, mat.category, mat.quantity, mat.unit, mat.manufacturer, mat.product,
        ], number_cols)

    _auto_column_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


EUR_FORMAT = '#,##0.00 "EUR"'


def generate_mengen(project: Project, quantities: list[QuantityItem]) -> io.BytesIO:
    """Generate Mengenermittlung (quantity takeoff) Excel report."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Mengenermittlung"

    ws.cell(row=1, column=1, value=f"Mengenermittlung - {project.name}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=project.address).font = SUBTITLE_FONT

    headers = ["Pos.", "Gewerk", "Kategorie", "Beschreibung", "Menge", "Einheit", "EP (EUR)", "GP (EUR)"]
    row = _write_header(ws, 4, headers)

    number_cols = {5, 7, 8}
    current_trade = None
    trade_total = 0.0
    pos = 0

    sorted_items = sorted(quantities, key=lambda q: (q.trade, q.category, q.element_type))

    for item in sorted_items:
        if current_trade is not None and item.trade != current_trade:
            row = _write_sum_row(ws, row, f"Summe {current_trade}", len(headers), 8, trade_total)
            ws.cell(row=row - 1, column=8).number_format = EUR_FORMAT
            trade_total = 0.0
            row += 1

        current_trade = item.trade
        trade_total += item.total_price
        pos += 1

        row = _write_row(ws, row, [
            pos, item.trade, item.category, f"{item.element_type} {item.description}".strip(),
            item.quantity, item.unit, item.unit_price, item.total_price,
        ], number_cols)

        ws.cell(row=row - 1, column=7).number_format = EUR_FORMAT
        ws.cell(row=row - 1, column=8).number_format = EUR_FORMAT

    if current_trade is not None:
        row = _write_sum_row(ws, row, f"Summe {current_trade}", len(headers), 8, trade_total)
        ws.cell(row=row - 1, column=8).number_format = EUR_FORMAT

    row += 1
    grand_total = sum(item.total_price for item in quantities)
    row = _write_sum_row(ws, row, "GESAMTSUMME", len(headers), 8, grand_total)
    ws.cell(row=row - 1, column=8).number_format = EUR_FORMAT

    _auto_column_width(ws)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
